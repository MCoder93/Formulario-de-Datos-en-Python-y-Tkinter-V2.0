# 🐺 Proyecto con Interfaz Gráfica en Python y Tkinter - V2.0.

# Este sistema es un formulario de entrada de datos que permite a los usuarios ingresar información,
# guardarla en un archivo Excel, visualizar los datos guardados y exportarlos a un archivo. 
# CSV. La interfaz gráfica está diseñada con Tkinter y cuenta con validaciones de entrada
# para asegurar que los datos ingresados sean correctos.

import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl 
from openpyxl import Workbook, load_workbook 
import os
import re
import csv

# Ruta del archivo Excel
archivo_excel = 'datos.xlsx'

# Verificar si el archivo Excel existe
if os.path.exists(archivo_excel):
    wb = load_workbook(archivo_excel)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Edad", "Email", "Teléfono", "Dirección"])
    wb.save(archivo_excel)

# Función para actualizar la barra de estado
def actualizar_estado(mensaje):
    estado_var.set(mensaje)

# Función para guardar los datos
def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefono = entry_telefono.get()
    direccion = entry_direccion.get()

    if not nombre or not edad or not email or not telefono or not direccion:
        messagebox.showwarning("Advertencia", "Todos los campos son obligatorios")
        actualizar_estado("⚠️ Campos incompletos")
        return

    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showwarning("Advertencia", "Edad y Teléfono deben ser números")
        actualizar_estado("⚠️ Edad o teléfono inválido")
        return

    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning("Advertencia", "Email no es válido")
        actualizar_estado("⚠️ Email inválido")
        return

    ws.append([nombre, edad, email, telefono, direccion])
    wb.save(archivo_excel)
    actualizar_estado("✅ Archivo guardado con éxito")

    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)

# Función para visualizar el archivo Excel
def visualizar_excel():
    os.startfile(archivo_excel)
    actualizar_estado("📂 Archivo Excel abierto")

# Función para exportar a CSV con selector de ruta
def exportar_a_csv():
    ruta = filedialog.asksaveasfilename(defaultextension=".csv",
                                        filetypes=[("Archivo CSV", "*.csv")],
                                        title="Guardar como")
    if ruta:
        with open(ruta, mode="w", newline="", encoding="utf-8") as archivo_csv:
            escritor = csv.writer(archivo_csv)
            escritor.writerow(["Nombre", "Edad", "Email", "Teléfono", "Dirección"])
            for fila in ws.iter_rows(min_row=2, values_only=True):
                escritor.writerow(fila)
        messagebox.showinfo("Exportación", f"Archivo exportado a:\n{ruta}")
        actualizar_estado("📤 Archivo exportado a CSV")

# Crear ventana principal
root = tk.Tk()
root.title("Formulario de Entrada de Datos")
root.configure(bg='#4B6587')
root.eval('tk::PlaceWindow . center')

label_style = {"bg": '#4B6587', "fg": "white", "font": ("Arial", 12)}
entry_style = {"bg": '#D3D3D3', "fg": "black", "font": ("Arial", 12)}

# Campos del formulario
tk.Label(root, text="Nombre", **label_style).grid(row=0, column=0, padx=20, pady=10)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=20, pady=10)

tk.Label(root, text="Edad", **label_style).grid(row=1, column=0, padx=20, pady=10)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=20, pady=10)

tk.Label(root, text="Email", **label_style).grid(row=2, column=0, padx=20, pady=10)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=20, pady=10)

tk.Label(root, text="Teléfono", **label_style).grid(row=3, column=0, padx=20, pady=10)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=20, pady=10)

tk.Label(root, text="Dirección", **label_style).grid(row=4, column=0, padx=20, pady=10)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=20, pady=10)

# Frame para agrupar los botones
boton_frame = tk.Frame(root, bg='#4B6587')
boton_frame.grid(row=5, column=0, columnspan=2, pady=20)

tk.Button(boton_frame, text="Guardar", command=guardar_datos,
            bg='#6D8299', fg='white', font=("Arial", 12), width=18).grid(row=0, column=0, padx=5)

tk.Button(boton_frame, text="Visualizar datos guardados", command=visualizar_excel,
            bg='#6D8299', fg='white', font=("Arial", 12), width=22).grid(row=0, column=1, padx=5)

tk.Button(boton_frame, text="Exportar archivo", command=exportar_a_csv,
            bg='#6D8299', fg='white', font=("Arial", 12), width=18).grid(row=0, column=2, padx=5)

# Barra de estado inferior
estado_var = tk.StringVar()
estado_var.set("📝 Listo para ingresar datos")
barra_estado = tk.Label(root, textvariable=estado_var, anchor="w",
                        bg="#2C3E50", fg="white", font=("Arial", 10))
barra_estado.grid(row=6, column=0, columnspan=2, sticky="we")

# Ajuste automático del tamaño de ventana
root.update_idletasks()  # Calcula tamaños
root.geometry(f"{root.winfo_width()}x{root.winfo_height()}")  # Aplica tamaño exacto

root.mainloop()


